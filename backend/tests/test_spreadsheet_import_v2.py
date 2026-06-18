from io import BytesIO
from openpyxl import Workbook

from app.services.spreadsheet_import_v2 import inspect_workbook_bytes


def make_template_like_workbook() -> bytes:
    wb = Workbook()
    wb.remove(wb.active)
    sheets = {
        "Products": ["product_id", "manufacturer", "category", "title", "status", "quality_level"],
        "Prices_Stock": ["product_id", "stock_status", "lead_time_days"],
        "Datasheet_Review": ["product_id", "field_name", "review_status", "corrected_value", "unit", "source_url"],
        "Labour_Rules": ["rule_id", "scope", "condition", "base_hours", "multiplier", "review_status"],
        "Mounting_Rules": ["mapping_id", "roof_type", "manufacturer", "system_family", "requires_manufacturer_calc", "review_status"],
        "Workflow_Feedback": ["feedback_id", "category", "severity", "description", "triage_status"],
    }
    for name, headers in sheets.items():
        ws = wb.create_sheet(name)
        ws.append(headers)
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


def test_valid_template_can_stage():
    report = inspect_workbook_bytes("template.xlsx", make_template_like_workbook())
    assert report.can_stage
    assert report.status == "staged"


def test_protected_engineering_field_blocks_products_sheet():
    data = make_template_like_workbook()
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("Products")
    ws.append(["product_id", "manufacturer", "category", "title", "status", "quality_level", "voc_v"])
    for name, headers in {
        "Prices_Stock": ["product_id", "stock_status", "lead_time_days"],
        "Datasheet_Review": ["product_id", "field_name", "review_status", "corrected_value", "unit", "source_url"],
        "Labour_Rules": ["rule_id", "scope", "condition", "base_hours", "multiplier", "review_status"],
        "Mounting_Rules": ["mapping_id", "roof_type", "manufacturer", "system_family", "requires_manufacturer_calc", "review_status"],
        "Workflow_Feedback": ["feedback_id", "category", "severity", "description", "triage_status"],
    }.items():
        w = wb.create_sheet(name); w.append(headers)
    bio = BytesIO(); wb.save(bio)
    report = inspect_workbook_bytes("bad.xlsx", bio.getvalue())
    assert not report.can_stage
    assert any(i.code == "PROTECTED_ENGINEERING_FIELDS_IN_EDITABLE_SHEET" for i in report.issues)
