from __future__ import annotations

from io import BytesIO
from pathlib import Path
import hashlib

from openpyxl import load_workbook

from app.schemas.spreadsheet import SheetInspection, SpreadsheetImportStatus, SpreadsheetInspectionReport
from app.schemas.validation import Severity, ValidationArea, ValidationIssue

REQUIRED_HEADERS: dict[str, set[str]] = {
    "Products": {"product_id", "manufacturer", "category", "title", "status", "quality_level"},
    "Prices_Stock": {"product_id", "stock_status", "lead_time_days"},
    "Datasheet_Review": {"product_id", "field_name", "review_status", "corrected_value", "unit", "source_url"},
    "Labour_Rules": {"rule_id", "scope", "condition", "base_hours", "multiplier", "review_status"},
    "Mounting_Rules": {"mapping_id", "roof_type", "manufacturer", "system_family", "requires_manufacturer_calc", "review_status"},
    "Workflow_Feedback": {"feedback_id", "category", "severity", "description", "triage_status"},
}

PROTECTED_ENGINEERING_HEADERS = {
    "voc_v", "isc_a", "vmp_v", "imp_a", "length_mm", "width_mm", "weight_kg",
    "max_dc_voltage_v", "mppt_min_v", "mppt_max_v", "wind_load", "ballast",
}


def sha256_bytes(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def _headers_from_sheet(ws) -> list[str]:
    values = []
    for cell in ws[1]:
        if cell.value is None:
            continue
        values.append(str(cell.value).strip())
    return values


def inspect_workbook_bytes(file_name: str, data: bytes) -> SpreadsheetInspectionReport:
    issues: list[ValidationIssue] = []
    try:
        wb = load_workbook(BytesIO(data), read_only=True, data_only=False)
    except Exception as exc:
        return SpreadsheetInspectionReport(
            file_name=file_name,
            file_hash_sha256=sha256_bytes(data),
            status=SpreadsheetImportStatus.failed_validation,
            issues=[ValidationIssue(
                code="WORKBOOK_OPEN_FAILED",
                severity=Severity.blocker,
                area=ValidationArea.spreadsheet,
                message=f"Workbook could not be opened: {exc}",
                suggested_fix="Upload a valid .xlsx file generated from the ArrayLab template.",
                blocks_status=True,
            )],
            can_stage=False,
        )

    inspections: list[SheetInspection] = []
    sheet_names = set(wb.sheetnames)
    for required_sheet, required_headers in REQUIRED_HEADERS.items():
        if required_sheet not in sheet_names:
            issues.append(ValidationIssue(
                code="REQUIRED_SHEET_MISSING",
                severity=Severity.error,
                area=ValidationArea.spreadsheet,
                message=f"Required sheet '{required_sheet}' is missing.",
                path=f"sheets.{required_sheet}",
                suggested_fix=f"Add the '{required_sheet}' sheet from the official template.",
            ))
            continue

        ws = wb[required_sheet]
        headers = _headers_from_sheet(ws)
        missing = sorted(required_headers - set(headers))
        if missing:
            issues.append(ValidationIssue(
                code="REQUIRED_HEADERS_MISSING",
                severity=Severity.error,
                area=ValidationArea.spreadsheet,
                message=f"Sheet '{required_sheet}' is missing required headers: {', '.join(missing)}",
                path=f"sheets.{required_sheet}.headers",
                suggested_fix="Use the official spreadsheet template or add the missing columns exactly.",
            ))

        protected_found = sorted(PROTECTED_ENGINEERING_HEADERS.intersection(headers))
        if protected_found and required_sheet not in {"Datasheet_Review"}:
            issues.append(ValidationIssue(
                code="PROTECTED_ENGINEERING_FIELDS_IN_EDITABLE_SHEET",
                severity=Severity.blocker,
                area=ValidationArea.spreadsheet,
                message=f"Sheet '{required_sheet}' includes protected engineering fields: {', '.join(protected_found)}",
                path=f"sheets.{required_sheet}.headers",
                suggested_fix="Move engineering spec edits into Datasheet_Review with source/provenance and reviewer status.",
                blocks_status=True,
            ))

        # Some generated workbooks report max_row/max_column as None in read-only mode.
        # Fall back to the header length so validation reports never crash on a valid workbook.
        row_count = int(ws.max_row or 1)
        column_count = int(ws.max_column or len(headers) or 0)
        inspections.append(SheetInspection(
            sheet_name=required_sheet,
            row_count=row_count,
            column_count=column_count,
            headers=headers,
            required_headers_missing=missing,
        ))

    status = SpreadsheetImportStatus.failed_validation if any(i.severity in {Severity.error, Severity.blocker} for i in issues) else SpreadsheetImportStatus.staged
    return SpreadsheetInspectionReport(
        file_name=file_name,
        file_hash_sha256=sha256_bytes(data),
        status=status,
        sheets=inspections,
        issues=issues,
        can_stage=status == SpreadsheetImportStatus.staged,
    )


def inspect_workbook_path(path: str | Path) -> SpreadsheetInspectionReport:
    p = Path(path)
    return inspect_workbook_bytes(p.name, p.read_bytes())
