from __future__ import annotations

from collections import defaultdict
from datetime import UTC, datetime
from io import BytesIO
from uuid import uuid4

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.db_models import ProductDataSnapshot, SpreadsheetImport, StagedImportRow
from app.schemas.spreadsheet import SpreadsheetImportStatus
from app.services.hash_utils import sha256_bytes, stable_json_hash
from app.services.spreadsheet_import_v2 import REQUIRED_HEADERS, inspect_workbook_bytes


class ImportWorkflowError(ValueError):
    """Raised when an import state transition would break governance rules."""


def _new_id(prefix: str) -> str:
    return f"{prefix}_{uuid4().hex[:14]}"


def _cell_value(value):
    if value is None:
        return None
    if isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def _extract_rows_by_sheet(data: bytes) -> list[dict]:
    """Extract non-empty spreadsheet rows into deterministic payloads.

    This deliberately ignores formula evaluation. `data_only=True` reads cached formula
    values if they exist, but engineering truth must still come from reviewed datasheets.
    """

    wb = load_workbook(BytesIO(data), read_only=True, data_only=True)
    staged: list[dict] = []
    for sheet_name in REQUIRED_HEADERS:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        rows = ws.iter_rows(values_only=True)
        try:
            headers = [str(h).strip() if h is not None else "" for h in next(rows)]
        except StopIteration:
            continue
        for row_number, row in enumerate(rows, start=2):
            values = [_cell_value(v) for v in row]
            if not any(v not in (None, "") for v in values):
                continue
            payload = {
                header: values[idx] if idx < len(values) else None
                for idx, header in enumerate(headers)
                if header
            }
            row_payload = {"sheet_name": sheet_name, "row_number": row_number, "payload": payload}
            staged.append(row_payload)
    return staged


def _diff_summary_for_rows(rows: list[dict], file_hash: str) -> dict:
    counts = defaultdict(int)
    hashes = defaultdict(list)
    for row in rows:
        counts[row["sheet_name"]] += 1
        hashes[row["sheet_name"]].append(stable_json_hash(row["payload"]))
    return {
        "mode": "first_class_staging_no_live_apply",
        "file_hash_sha256": file_hash,
        "total_staged_rows": len(rows),
        "rows_by_sheet": dict(sorted(counts.items())),
        "sheet_hashes": {k: stable_json_hash(v) for k, v in sorted(hashes.items())},
        "note": "Approval creates an immutable product-data snapshot. It does not mutate old projects or live engineering specs.",
    }


def stage_spreadsheet_import(db: Session, file_name: str, data: bytes, uploaded_by: str | None = None) -> SpreadsheetImport:
    inspection = inspect_workbook_bytes(file_name, data)
    file_hash = sha256_bytes(data)
    rows = _extract_rows_by_sheet(data) if inspection.can_stage else []
    diff_summary = _diff_summary_for_rows(rows, file_hash) if inspection.can_stage else {
        "mode": "failed_validation_no_staged_rows",
        "file_hash_sha256": file_hash,
        "total_staged_rows": 0,
    }

    record = SpreadsheetImport(
        import_id=_new_id("imp"),
        file_name=file_name,
        file_hash_sha256=file_hash,
        status=inspection.status.value,
        uploaded_by=uploaded_by,
        validation_report=inspection.model_dump(mode="json"),
        diff_summary=diff_summary,
    )
    db.add(record)
    db.flush()

    for row in rows:
        payload = row["payload"]
        row_hash = stable_json_hash({"sheet_name": row["sheet_name"], "row_number": row["row_number"], "payload": payload})
        db.add(StagedImportRow(
            staged_row_id=_new_id("sir"),
            import_id=record.import_id,
            sheet_name=row["sheet_name"],
            row_number=int(row["row_number"]),
            row_hash_sha256=row_hash,
            payload=payload,
            status="pending",
        ))

    db.commit()
    db.refresh(record)
    return record


def get_spreadsheet_import(db: Session, import_id: str) -> SpreadsheetImport | None:
    return db.get(SpreadsheetImport, import_id)


def list_spreadsheet_imports(db: Session) -> list[SpreadsheetImport]:
    return list(db.scalars(select(SpreadsheetImport).order_by(SpreadsheetImport.created_at.desc())))


def list_staged_rows(db: Session, import_id: str) -> list[StagedImportRow]:
    return list(db.scalars(
        select(StagedImportRow)
        .where(StagedImportRow.import_id == import_id)
        .order_by(StagedImportRow.sheet_name, StagedImportRow.row_number)
    ))


def approve_staged_import(db: Session, import_id: str, approved_by: str) -> ProductDataSnapshot:
    record = get_spreadsheet_import(db, import_id)
    if record is None:
        raise ImportWorkflowError(f"Import {import_id} not found.")
    if record.status != SpreadsheetImportStatus.staged.value:
        raise ImportWorkflowError(f"Only staged imports can be approved. Current status is {record.status}.")

    rows = list_staged_rows(db, import_id)
    if not rows:
        raise ImportWorkflowError("Cannot approve import with zero staged rows.")

    snapshot_payload = {
        "source_import_id": import_id,
        "source_file_hash_sha256": record.file_hash_sha256,
        "rows": [
            {
                "sheet_name": row.sheet_name,
                "row_number": row.row_number,
                "row_hash_sha256": row.row_hash_sha256,
                "payload": row.payload,
            }
            for row in rows
        ],
    }
    content_hash = stable_json_hash(snapshot_payload)
    snapshot = ProductDataSnapshot(
        snapshot_id=f"pds_{content_hash[:16]}",
        import_id=import_id,
        content_hash_sha256=content_hash,
        row_count=len(rows),
        summary={
            "approved_from_import": import_id,
            "approved_by": approved_by,
            "rows_by_sheet": record.diff_summary.get("rows_by_sheet", {}),
            "truth_boundary": "snapshot only; no old project mutation",
        },
        snapshot_payload=snapshot_payload,
        created_by=approved_by,
    )
    db.add(snapshot)
    record.status = SpreadsheetImportStatus.approved.value
    record.approved_by = approved_by
    record.approved_at = datetime.now(UTC)
    db.commit()
    db.refresh(snapshot)
    return snapshot


def reject_staged_import(db: Session, import_id: str, rejected_by: str, reason: str | None = None) -> SpreadsheetImport:
    record = get_spreadsheet_import(db, import_id)
    if record is None:
        raise ImportWorkflowError(f"Import {import_id} not found.")
    if record.status not in {SpreadsheetImportStatus.staged.value, SpreadsheetImportStatus.failed_validation.value, SpreadsheetImportStatus.uploaded.value}:
        raise ImportWorkflowError(f"Import in status {record.status} cannot be rejected.")
    record.status = SpreadsheetImportStatus.rejected.value
    record.rejected_by = rejected_by
    record.rejected_at = datetime.now(UTC)
    record.diff_summary = {**(record.diff_summary or {}), "rejection_reason": reason}
    db.commit()
    db.refresh(record)
    return record


def latest_product_data_snapshot(db: Session) -> ProductDataSnapshot | None:
    return db.scalars(select(ProductDataSnapshot).order_by(ProductDataSnapshot.created_at.desc())).first()
