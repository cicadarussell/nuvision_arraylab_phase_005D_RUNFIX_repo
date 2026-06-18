from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook

ROOT = Path(__file__).resolve().parents[1]
SEED = ROOT / "data" / "seed" / "nuvision_product_catalog_seed_v0.json"
OUT = ROOT / "data" / "seed" / "nuvision_seed_products_for_review.xlsx"


def main() -> int:
    data = json.loads(SEED.read_text(encoding="utf-8"))
    wb = Workbook()
    wb.remove(wb.active)

    products = wb.create_sheet("Products")
    products.append(["product_id", "manufacturer", "category", "title", "status", "quality_level", "nuvision_sku", "manufacturer_model", "nuvision_url"])

    review = wb.create_sheet("Datasheet_Review")
    review.append(["product_id", "field_name", "review_status", "corrected_value", "unit", "source_url"])

    for idx, item in enumerate(data.get("products", {}).get("solar_pv_panels_public_collection", []), start=1):
        product_id = f"seed_panel_{idx:03d}"
        products.append([
            product_id,
            item.get("brand") or "unknown",
            "panel" if item.get("power_w") else "other",
            item.get("title") or f"Seed product {idx}",
            "active" if str(item.get("availability_text", "")).lower().find("stock") >= 0 else "unknown",
            "Q0_scraped",
            item.get("sku"),
            item.get("model"),
            item.get("url"),
        ])
        if item.get("power_w"):
            review.append([product_id, "power_stc_w", "unreviewed", item.get("power_w"), "W", item.get("url")])

    # Required empty sheets for controlled import template compatibility.
    for name, headers in {
        "Prices_Stock": ["product_id", "stock_status", "lead_time_days"],
        "Labour_Rules": ["rule_id", "scope", "condition", "base_hours", "multiplier", "review_status"],
        "Mounting_Rules": ["mapping_id", "roof_type", "manufacturer", "system_family", "requires_manufacturer_calc", "review_status"],
        "Workflow_Feedback": ["feedback_id", "category", "severity", "description", "triage_status"],
    }.items():
        ws = wb.create_sheet(name)
        ws.append(headers)

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Wrote {OUT}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
