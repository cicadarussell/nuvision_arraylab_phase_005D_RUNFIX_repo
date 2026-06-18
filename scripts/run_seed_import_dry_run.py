from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import Workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

ROOT = Path(__file__).resolve().parents[1]
BACKEND = ROOT / "backend"
sys.path.insert(0, str(BACKEND))

from app.models.db_models import Base
from app.services.product_snapshot_apply import build_product_apply_preview
from app.services.spreadsheet_staging import approve_staged_import, stage_spreadsheet_import

SEED = ROOT / "data" / "seed" / "nuvision_product_catalog_seed_v0.json"
OUT = ROOT / "data" / "seed" / "nuvision_seed_import_dry_run_report.json"


def make_seed_workbook() -> bytes:
    data = json.loads(SEED.read_text(encoding="utf-8"))
    wb = Workbook()
    wb.remove(wb.active)

    products = wb.create_sheet("Products")
    products.append(["product_id", "manufacturer", "category", "title", "status", "quality_level", "nuvision_sku", "manufacturer_model", "nuvision_url"])
    review = wb.create_sheet("Datasheet_Review")
    review.append(["product_id", "field_name", "review_status", "corrected_value", "unit", "source_url"])
    price_stock = wb.create_sheet("Prices_Stock")
    price_stock.append(["product_id", "stock_status", "lead_time_days", "trade_price_gbp", "list_price_gbp", "currency", "stock_quantity", "supplier_priority"])

    panel_items = data.get("products", {}).get("solar_pv_panels_public_collection", [])
    for idx, item in enumerate(panel_items, start=1):
        product_id = f"seed_panel_{idx:03d}"
        products.append([
            product_id,
            item.get("brand") or "unknown",
            "panel",
            item.get("title") or f"Seed panel {idx}",
            "active" if "stock" in str(item.get("availability_text", "")).lower() else "hidden",
            "Q0_scraped",
            item.get("sku"),
            item.get("model"),
            item.get("url"),
        ])
        if item.get("power_w"):
            review.append([product_id, "power_stc_w", "unreviewed", item.get("power_w"), "W", item.get("url")])
        price_stock.append([product_id, "unknown", None, None, None, "GBP", None, None])

    for name, headers in {
        "Labour_Rules": ["rule_id", "scope", "condition", "base_hours", "multiplier", "review_status"],
        "Mounting_Rules": ["mapping_id", "roof_type", "manufacturer", "system_family", "requires_manufacturer_calc", "review_status"],
        "Workflow_Feedback": ["feedback_id", "category", "severity", "description", "triage_status"],
    }.items():
        ws = wb.create_sheet(name)
        ws.append(headers)

    import io
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def main() -> int:
    engine = create_engine("sqlite+pysqlite:///:memory:", future=True)
    Base.metadata.create_all(engine)
    SessionLocal = sessionmaker(bind=engine, future=True)
    with SessionLocal() as db:
        workbook = make_seed_workbook()
        record = stage_spreadsheet_import(db, "nuvision_seed_import_dry_run.xlsx", workbook, uploaded_by="seed_dry_run")
        result = {
            "import_status": record.status,
            "total_staged_rows": record.diff_summary.get("total_staged_rows"),
            "rows_by_sheet": record.diff_summary.get("rows_by_sheet", {}),
            "issues": record.validation_report.get("issues", []),
        }
        if record.status == "staged":
            snapshot = approve_staged_import(db, record.import_id, approved_by="seed_dry_run")
            app = build_product_apply_preview(db, snapshot.snapshot_id, created_by="seed_dry_run")
            result.update({
                "snapshot_id": snapshot.snapshot_id,
                "preview_status": app.status,
                "preview_actions": app.diff_summary.get("actions", {}),
                "preview_warning_count": sum(1 for item in app.diff_summary.get("diff_items", []) if item.get("warnings")),
                "truth_boundary": "seed products stay Q0 until manufacturer datasheet review; dry-run does not apply to live product table",
            })
        OUT.parent.mkdir(parents=True, exist_ok=True)
        OUT.write_text(json.dumps(result, indent=2), encoding="utf-8")
        print(json.dumps(result, indent=2))
        return 0 if result["import_status"] == "staged" else 1


if __name__ == "__main__":
    raise SystemExit(main())
